import pandas as pd
import time
import operator
import subprocess
import openpyxl
import os

def get_ping_list(filename, sheet,tables): #从excel表里读取objec列的所有数值
    table_device = pd.read_excel(filename,sheet)
    ping_list = table_device[tables]
    return  ping_list

def get_change_ping_list(ping_list):#序列化exce表的ip,让程序能能够处理
    change_ping_list =[]
    for ip in ping_list:
        change_ping_list.append(ip)
    # print("resutl_list",result_list)
    fping_ping_list = " ".join(change_ping_list) #数组增加空格符合fping格式
    return fping_ping_list

def masscan_run(fping_ping_list): #masscan扫描并返回发现的ip列表
    m = subprocess.getoutput('masscan -p 22 %s --rate=100000'%fping_ping_list) 
    m_list_ip= []
    m_list = m.split('\n')
    for m_arr in m_list:
        if 'Discovered' in m_arr:
            m_arr_list = m_arr.split(" ")
            m_list_ip.append(m_arr_list[5])#将IP放入
    return m_list_ip

def fping_run(fping_ping_list,m_list_ip):#fping检测，并返回存活值
    p = subprocess.getoutput('fping -a -c 3 %s'%fping_ping_list) 
    device_live_list=[]
    fping_list = p.split('\n') #进行拆分结果
    for fping_arr in  fping_list:
        if 'xmt' in fping_arr:
            fping_ip_init = fping_arr.split(":")
            fping_ip = fping_ip_init[0].strip()#ip两边空格去除
            device_live_init = False
            for m_ip in m_list_ip:
                if fping_ip == m_ip:
                    device_live_init = True
            device_live = operator.contains(fping_arr,"min") or device_live_init #fping检测或者masscan扫描有其中一项即可
            device_live_list.append(change_result(device_live)) #将True转换为1
    return device_live_list


def write_to_excel(filename, write_filename,sheet, ncols, write_result):#写入excel并保存
    wb = pd.read_excel(filename,sheet)
    #wb.insert(loc=ncols, column=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),value=write_result, allow_duplicates=True)#新增一列并插入相应的数值
    wb.insert(loc=ncols, column='在线情况',value=write_result, allow_duplicates=True)#新增一列并插入相应的数值
    wb.to_excel(write_filename)

def change_result(result): #将结果转换为1或者0,看你喜欢
    if result:
        return 1 
    else:
        return 0 

def delete_clo(excel_name,excel_delete_name,excel_sheet_work):
    wb= openpyxl.load_workbook(excel_name)
    ws = wb[excel_sheet_work]
    ws.delete_cols(1)
    #saveexcel = excel_name.strip('.xlsx')+str(round(time.time()))+'.xlsx'#加入时间戳
    wb.save(excel_delete_name)      


def get_work_value(excel_name):  #筛选工作界面，得到索引和工作界面值
    df =pd.read_excel(excel_name)#读取excel
    listindex = [] #索引数组
    listpole = [] #杆号
    listallresult = []
    for index,row in df.iterrows():                   #循环
        listindex.append(row[0])                      #获取index
        listpole.append(row[8])                      #获取杆子号
        listindex_sort =list(set(listindex))        #进行排序从小到大并去重
        listpole_sort = list(set(listpole))         #进行排序从小到大并去重
    for index_loop in listindex_sort:   #序号遍历
        for pole_loop in listpole_sort: #杆子号遍历
            pole_result = df[(df.序号==index_loop)& (df.百度杆位==pole_loop)] #先遍历路口,然后遍历每个路口的杆子
            if not pole_result.empty: #过滤掉路口没有这个杆号
                pole_result_in = pole_result.在线情况.isin([1]).any() #杆号存活率里面筛选出包含1的数值，进行判断
                pole_result_index = pole_result.index #返回某个路口，某一杆子的索引

                if pole_result_in: #如果包含1
                    for  b_index in pole_result_index:
                        listallresult.append(str(b_index) + ' '+'类型A')
                else:   #如果全部为0，那么为0
                    for w_index in pole_result_index:
                        listallresult.append(str(w_index)+' '+ '类型B')
    return listallresult

def write_excel(excel_name,excel_sheet,list_workvalue,execl_ncols):
    wb= openpyxl.load_workbook(excel_name)
    ws = wb[excel_sheet]
    #ws.delete_cols(1)
    for workvalue in list_workvalue:
        ws.cell(row=int(workvalue.split(' ')[0])+2,column=execl_ncols).value=workvalue.split(' ')[1]
    #saveexcel = excel_name.strip('.xlsx')+str(round(time.time()))+'.xlsx'#加入时间戳
    #saveexcel = 'YZ2V2X'+str(round(time.time()))+'.xlsx'#加入时间戳
    wb.save('yzpzb3.xlsx')

excel_name = './text.xlsx' #源表
excel_sheet = 'Sheet2'
excel_table = 'IP'
#execl_ncols = 15 #插入表的第几列
execl_ncols = 15 #插入表的第几列
execl_writename ='./test1.xlsx'#保存表的名字

excel_delelt = './test2.xlsx'
excel_sheet_work = 'Sheet1'
execl_ncols_work = 17 #插入表的第几列

ping_list = get_ping_list(excel_name, excel_sheet,excel_table) #得到excel的某一列的数值
fping_ping_list = get_change_ping_list(ping_list)   #序列化能让fping执行的数组
m_list_ip = masscan_run(fping_ping_list)            #得到mascan扫描的结果IP组
device_live_list=fping_run(fping_ping_list,m_list_ip)   #通过fping结合masscan的结果得到存活率

write_to_excel(excel_name,execl_writename, excel_sheet, execl_ncols, device_live_list) #写入excel先筛选

delete_clo(execl_writename,excel_delelt,excel_sheet_work) #删除生成excel的第一列
list_allresult = get_work_value(excel_delelt) #得到所有的排序结果
write_excel(excel_delelt,excel_sheet_work,list_allresult,execl_ncols_work)

os.remove(execl_writename)
os.remove(excel_delelt)
