import pandas as pd
import time
import operator
import subprocess
import openpyxl
import os
import numpy as np

cameraCount = 0
cameraLossCount =0 
cameraLossCheck =0
cameraLossNoCheck= 0
cameraListLossAll = []

ccuCount = 0
ccuLossCount =0 
ccuLossCheck =0
ccuLossNoCheck= 0
ccuListLossAll = []

rsuCount = 0
rsuLossCount =0 
rsuLossCheck =0
rsuLossNoCheck= 0
rsuListLossAll = []

radarCount = 0
radarLossCount =0 
radarLossCheck =0
radarLossNoCheck= 0
radarListLossAll = []

def get_camera_value(excel_name):  #筛选工作界面，得到索引和工作界面值
    # global countNum,lossCountNum,lossCheckNum,lossNoCheckNum
    df =pd.read_excel(excel_name)#读取excel
    listAll = []
    listLossAll = []
    n = 1
    lossCount = 0
    lossCheck = 0
    clos =[i for i in df.columns if i not in ['Time']]
    df2 = df[clos]
    for column in df2.columns:
        n= n+1 
        listallresult = df2[column].tolist()
        a = np.array(df2[column].tolist())
        vu = a[(np.where((a>0) & (a<1000)))]
        b1,s1 = np.unique(vu,return_counts=True)
        # print("column",column)
        #b,s ,t,w= np.unique(a,return_counts=True,return_index=True,return_inverse=True)
        sumNum = int(np.nansum(a)) #求和把nan去掉
        # print("n",n,"sumNum",sumNum)
        if vu.sum() > 90:
            lossCount = lossCount+1
            columnV = str(n) + ' ' + str(1) + ' ' + column + ' ' +str(sumNum)
            if s1.size >= 5:
                lossCheck = lossCheck+1
                listAll.append(columnV)
                listLossAll.append(column)
            if s1.size == 2 and s1.sum() >= 26 :
                lossCheck = lossCheck+1
                listAll.append(columnV)
                listLossAll.append(column)
            if s1.size == 3 and s1.sum() >= 14:
                lossCheck = lossCheck+1
                listAll.append(columnV)
                listLossAll.append(column)
            if s1.size == 4 and s1.sum() >= 4:
                lossCheck = lossCheck+1
                listAll.append(columnV)
                listLossAll.append(column)
        else:
            columnV = str(n) + ' ' + str(0) + ' ' + column + ' ' +str(sumNum)
            listAll.append(columnV)
    # print("n",n)
    #print(listAll)
    countNum = n -1
    lossCountNum = lossCount
    lossCheckNum = lossCheck
    lossNoCheckNum=  lossCountNum - lossCheckNum
    # print("countNum",countNum)
    # print("lossCountNum",lossCountNum)
    # print("lossCheckNum",lossCheckNum)
    # print("lossNoCheckNum",lossNoCheckNum)
    # print("listLossAll",listLossAll)
    return listAll,countNum,lossCountNum,lossCheckNum,lossNoCheckNum,listLossAll

def write_excel(excel_name,excel_sheet,excel_save,list_value):
    wb= openpyxl.load_workbook(excel_name)
    ws = wb[excel_sheet]
    #ws.delete_cols(1)
    for cameravalue  in list_value:
        ws.cell(row=291,column=int(cameravalue.split(' ')[0])).value=int(cameravalue.split(' ')[1])
  
    wb.save(excel_save)

def write_excel1(excel_name,excel_sheet,excel_save,list_value):
    wb= openpyxl.load_workbook(excel_name)
    ws1 = wb[excel_sheet]
    #ws.delete_cols(1)
    # for cameravalue  in list_value:
    #     ws.cell(row=291,column=int(cameravalue.split(' ')[0])).value=int(cameravalue.split(' ')[1])
  
    ws1.cell(row=1,column=1).value='路口号'
    ws1.cell(row=1,column=2).value='设备号'
    ws1.cell(row=1,column=3).value='IP'
    ws1.cell(row=1,column=4).value='丢包数'
    ws1.cell(row=1,column=5).value='是否排查'
    for i,valueR in enumerate(list_value):
        # print(valueR)
        roleNume = valueR.split('-')[1]
        deviceIP = valueR.split('-')[2].split('_')[1].split(' ')[0]
        deviceName = valueR.split(' ')[2]
        checkReult = valueR.split(' ')[1]
        pingLoss =valueR.split(' ')[3]

        ws1.cell(row=i+2,column=1).value=int(roleNume)
        ws1.cell(row=i+2,column=2).value=deviceName
        ws1.cell(row=i+2,column=3).value=deviceIP
        ws1.cell(row=i+2,column=4).value=int(pingLoss)
        ws1.cell(row=i+2,column=5).value=int(checkReult)

    wb.save(excel_save)


def get_normal_value(excel_name):  #筛选工作界面，得到索引和工作界面值
    df =pd.read_excel(excel_name)#读取excel
    listAll = []
    listLossAll = []
    n = 1
    lossCount = 0
    lossCheck = 0
    clos =[i for i in df.columns if i not in ['Time']]
    df2 = df[clos]
    for column in df2.columns:
        n= n+1 
        listallresult = df2[column].tolist()
        a = np.array(df2[column].tolist())
        vu = a[(np.where((a>0) & (a<1000)))]
        b1,s1 = np.unique(vu,return_counts=True)
        # print("column",column)
        #b,s ,t,w= np.unique(a,return_counts=True,return_index=True,return_inverse=True)
        sumNum = int(np.nansum(a)) #求和把nan去掉
        # print("n",n,"sumNum",sumNum)
        if vu.sum() > 30:
            lossCount = lossCount+1
            columnV = str(n) + ' ' + str(1) + ' ' + column + ' ' +str(sumNum)
            if s1.size >= 5:
                lossCheck = lossCheck+1
                listAll.append(columnV)
                listLossAll.append(column)
            if s1.size == 2 and s1.sum() >= 26 :
                lossCheck = lossCheck+1
                listAll.append(columnV)
                listLossAll.append(column)
            if s1.size == 3 and s1.sum() >= 14:
                lossCheck = lossCheck+1
                listAll.append(columnV)
                listLossAll.append(column)
            if s1.size == 4 and s1.sum() >= 4:
                lossCheck = lossCheck+1
                listAll.append(columnV)
                listLossAll.append(column)
        else:
            columnV = str(n) + ' ' + str(0) + ' ' + column + ' ' +str(sumNum)
            listAll.append(columnV)
    # print("n",n)
    #print(listAll)
    countNum = n -1
    lossCountNum = lossCount
    lossCheckNum = lossCheck
    lossNoCheckNum=  lossCountNum - lossCheckNum
    # print("countNum",countNum)
    # print("lossCountNum",lossCountNum)
    # print("lossCheckNum",lossCheckNum)
    # print("lossNoCheckNum",lossNoCheckNum)
    # print("listLossAll",listLossAll)
    return listAll,countNum,lossCountNum,lossCheckNum,lossNoCheckNum,listLossAll


def impact_road(listLossAll):
    roadList = [2,9,55,60,62,64,74,75,76,80,81,83,93,97,98,100,101,102,104,107,119,122,129,135,136,141,142,148,157,182,184,188,189,192,205,209,214,215,216,217,238,240,242,244,249,267,268,282,286,295,299]
    impactRoadNum=0
    roadLossList = []
    for i,valueR in enumerate(listLossAll):
        roadNum = int(valueR.split('-')[1])
        roadLossList.append(roadNum)
        # print("roadNum",roadNum)

    roadLossListSort = list(set(roadLossList)) 
    # print("roadLossList",roadLossList)
    # print("roadLossListSort",roadLossListSort)
    # print("set(roadLossList)",set(roadLossList))
    for roadLossNum in roadLossListSort:
        if roadLossNum in roadList:
            impactRoadNum = impactRoadNum+1

    return impactRoadNum
    # print("listall",listallresult[7])    
excel_camera_name = 'camera.xlsx' #源表
excel_camera_sheet = 'RSCU到感知相机通信-data-as-seriestocol'
excel_camera_save = 'cameracheck1.xlsx'
excel_camera_sheet1 = 'Sheet1'
excel_camera_save1 = 'cameracheck.xlsx'

excel_ccu_name = 'ccu.xlsx' #源表
excel_ccu_sheet = 'RSCU到CCU丢包率-data-as-seriestocol'
excel_ccu_save = 'ccucheck1.xlsx'
excel_ccu_sheet1 = 'Sheet1'
excel_ccu_save1 = 'ccucheck.xlsx'

excel_rsu_name = 'rsu.xlsx' #源表
excel_rsu_sheet = 'RSCU到RSU丢包率-data-as-seriestocol'
excel_rsu_save = 'rsucheck1.xlsx'
excel_rsu_sheet1 = 'Sheet1'
excel_rsu_save1 = 'rsucheck.xlsx'

excel_radar_name = 'radar.xlsx' #源表
excel_radar_sheet = 'RSCU到感知雷达通信-data-as-seriestocol'
excel_radar_save = 'radarcheck1.xlsx'
excel_radar_sheet1 = 'Sheet1'
excel_radar_save1 = 'radarcheck.xlsx'


camerList,cameraCount,cameraLossCount,cameraLossCheck,cameraLossNoCheck,cameraListLossAll = get_camera_value(excel_camera_name)
cameraImpactRoad= impact_road(cameraListLossAll)
write_excel(excel_camera_name,excel_camera_sheet,excel_camera_save,camerList)
write_excel1(excel_camera_save,excel_camera_sheet1,excel_camera_save1,camerList)

ccuList,ccuCount,ccuLossCount,ccuLossCheck,ccuLossNoCheck,ccuListLossAll = get_normal_value(excel_ccu_name)
ccuImpactRoad= impact_road(ccuListLossAll)
write_excel(excel_ccu_name,excel_ccu_sheet,excel_ccu_save,ccuList)
write_excel1(excel_ccu_save,excel_ccu_sheet1,excel_ccu_save1,ccuList)

rsuList,rsuCount,rsuLossCount,rsuLossCheck,rsuLossNoCheck,rsuListLossAll = get_normal_value(excel_rsu_name)
rsuImpactRoad= impact_road(rsuListLossAll)
write_excel(excel_rsu_name,excel_rsu_sheet,excel_rsu_save,rsuList)
write_excel1(excel_rsu_save,excel_rsu_sheet1,excel_rsu_save1,rsuList)

radarList,radarCount,radarLossCount,radarLossCheck,radarLossNoCheck,radarListLossAll = get_normal_value(excel_radar_name)
radarImpactRoad= impact_road(radarListLossAll)
write_excel(excel_radar_name,excel_radar_sheet,excel_radar_save,radarList)
write_excel1(excel_radar_save,excel_radar_sheet1,excel_radar_save1,radarList)

countSum = cameraCount + ccuCount + rsuCount + radarCount
lossCountSum = cameraLossCount + ccuLossCount + rsuLossCount + radarLossCount
lossNoCheckSum = cameraLossNoCheck + ccuLossNoCheck + rsuLossNoCheck + radarLossNoCheck
lossCheckSum = cameraLossCheck + ccuLossCheck + rsuLossCheck + radarLossCheck
impactRoadSum = cameraImpactRoad + ccuImpactRoad + rsuImpactRoad + radarImpactRoad

allRustlList = pd.DataFrame({'总设备数量':[cameraCount,ccuCount,rsuCount,radarCount,countSum],
'丢包设备总数':[cameraLossCount,ccuLossCount,rsuLossCount,radarLossCount,lossCountSum],
'丢包不排查':[cameraLossNoCheck,ccuLossNoCheck,rsuLossNoCheck,radarLossNoCheck,lossNoCheckSum],
'丢包排查':[cameraLossCheck,ccuLossCheck,rsuLossCheck,radarLossCheck,lossCheckSum],
'涉及重点路口数量':[cameraImpactRoad,ccuImpactRoad,rsuImpactRoad,radarImpactRoad,impactRoadSum]},index = ["相机","串口设备","RSU","雷达","共计"])
# print(allRustlList)
allRustlList.to_excel('分析结果.xlsx')

os.remove(excel_camera_save)
os.remove(excel_ccu_save)
os.remove(excel_rsu_save)
os.remove(excel_radar_save)